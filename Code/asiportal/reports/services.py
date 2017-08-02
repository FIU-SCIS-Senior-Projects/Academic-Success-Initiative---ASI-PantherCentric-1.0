from openpyxl.chart import (
        PieChart,
        Reference,
        Series,
        BarChart
        )
import datetime
from collections import OrderedDict
from django.db.models import Avg
from django.contrib.auth.models import User
from django.utils import timezone
from courses.models import Course
from rquests.models import TutoringRequest
from tutoring_sessions.models import Session, IndividualSession
from surveys.models import AmbassadorSurvey, TuteeSurvey
from timesheets.models import TutoringTimeSheetEntry, ProjectTimeSheetEntry

def prepare_range(semester):
    rightnow = timezone.now().date()
    start_of_week = semester.start_date
    diff = rightnow - start_of_week
    weeks_since = diff.days // 7 # int division
    return (weeks_since, start_of_week)

def get_overall_weeks_stats(view, semester):
    tr = TutoringRequest.objects.all() # this is quicker
    ambs = AmbassadorSurvey.objects.all()
    tuts = TuteeSurvey.objects.all()
    sess = Session.objects.all()
    weeks_since, start_of_week = prepare_range(semester)
    weeks = OrderedDict()
    for n in range(weeks_since+1): # +1 to include the week we're on right now
        date_range = (start_of_week, start_of_week + datetime.timedelta(days=6))
        requests_of_this_week = tr.filter(
                submitted_at__range=date_range)
        ambs_this_week = ambs.filter(
                individual_session__session_date__range=date_range)
        tuts_this_week = tuts.filter(
                individual_session__session_date__range=date_range)
        stats = OrderedDict()
        week_range = '{:%m/%d/%y} - {:%m/%d/%y}'.format(*date_range)
        stats['Total Tutoring Requests'] = requests_of_this_week.count()
        stats['Total Tutoring Requests Compelted'] = requests_of_this_week.filter(
                updated_at__range=date_range).count()
        stats['Total Tutoring Requests Compelted'] = requests_of_this_week.filter(
                updated_at__range=date_range).count()
        stats['Total Sessions Scheduled For This Week'] = sess.filter(
                start_date__lte=date_range[1],
                ).count()
        stats['Total Ambassador Surveys Expected'] = ambs_this_week.count()
        stats['Total Tutee Surveys Expected'] = tuts_this_week.count()
        stats['Total Ambassador Surveys Completed'] = ambs_this_week.filter(
                submitted=True).count()
        stats['Total Tutee Surveys Completed'] = tuts_this_week.filter(
                submitted=True).count()
        weeks[week_range] = stats
        start_of_week = start_of_week + datetime.timedelta(days=7)
    return weeks

def get_course_week_stats(view, semester):
    c = Course.objects.all()
    tr = TutoringRequest.objects.all()
    weeks = OrderedDict()
    weeks_since, start_of_week = prepare_range(semester)
    for n in range(weeks_since+1): 
        date_range = (start_of_week, start_of_week + datetime.timedelta(days=6))
        week_range = '{:%m/%d/%y} - {:%m/%d/%y}'.format(*date_range)
        stats = dict()
        for course in c:
            stats[course] = tr.filter(course=course, 
                    submitted_at__range=date_range,
                    ).count()
            weeks[week_range] = OrderedDict(sorted(stats.items(), key=lambda x : x[1], reverse=True))
        start_of_week = start_of_week + datetime.timedelta(days=7)
    return weeks

def get_session_per_course_week_stats(view, semester):
    c = Course.objects.all()
    weeks = OrderedDict()
    weeks_since, start_of_week = prepare_range(semester)
    for n in range(weeks_since+1): 
        date_range = (start_of_week, start_of_week + datetime.timedelta(days=6))
        week_range = '{:%m/%d/%y} - {:%m/%d/%y}'.format(*date_range)
        stats = dict()
        for course in c:
            se = course.course_sessions.filter(
                    start_date__lte=date_range[1],
                    end_date__gt=date_range[1])
            stats[course] = se.count()
            weeks[week_range] = OrderedDict(sorted(stats.items(),
                key=lambda x : x[1], reverse=True))
        start_of_week = start_of_week + datetime.timedelta(days=7)
    return weeks

def get_survey_per_week_stats(view, semester):
    ambs = AmbassadorSurvey.objects.all()
    tuts = TuteeSurvey.objects.all()
    weeks = OrderedDict()
    weeks_since, start_of_week = prepare_range(semester)
    for n in range(weeks_since+1): 
        date_range = (start_of_week, start_of_week + datetime.timedelta(days=6))
        week_range = '{:%m/%d/%y} - {:%m/%d/%y}'.format(*date_range)
        stats = OrderedDict()
        expected_ambs = ambs.filter(
                individual_session__session_date__range=date_range,
                ).count()
        expected_tuts = tuts.filter(
                individual_session__session_date__range=date_range,
                    ).count()
        submitted_ambs = ambs.filter(
                individual_session__session_date__range=date_range,
                submitted=True,
                ).count()
        submitted_tuts = tuts.filter(
                individual_session__session_date__range=date_range,
                submitted=True,
                ).count()
        stats['Expected Ambassador Surveys'] = expected_ambs
        stats['Expected Tutee Surveys'] = expected_tuts
        stats['Submitted Ambassador Surveys'] = submitted_ambs
        stats['Submitted Tutee Surveys'] = submitted_tuts
        stats['Total Expected Surveys'] = expected_ambs + expected_tuts
        stats['Total Actual Submitted Surveys'] = submitted_tuts + submitted_ambs
        stats['Percent Ambassador Survey Completed'] = submitted_ambs/expected_ambs * 100 if expected_ambs > 0 else "No Surveys Expected"
        stats['Percent Tutee Survey Completed'] = submitted_tuts/expected_tuts * 100 if expected_tuts > 0 else "No Surveys Expected"
        weeks[week_range] = stats
        start_of_week = start_of_week + datetime.timedelta(days=7)
    return weeks

def get_ambassador_stats(ambassador, semester):
    weeks = OrderedDict()
    weeks_since, start_of_week = prepare_range(semester)
    sessions = Session.objects.filter(availability__ambassador=ambassador)
    surveys = AmbassadorSurvey.objects.filter(
            individual_session__session__availability__ambassador=ambassador)
    for n in range(weeks_since+1): 
        stats = OrderedDict()
        date_range = (start_of_week, start_of_week + datetime.timedelta(days=6))
        week_range = '{:%m/%d/%y} - {:%m/%d/%y}'.format(*date_range)
        stats['Total Scheduled Sessions'] = sessions.filter(
                start_date__range=date_range,
                ).count()
        stats['Total Surveys Expected'] = surveys.filter(
                individual_session__session_date__range=date_range,
                ).count()
        stats['Total Surveys Submitted'] = surveys.filter(
                individual_session__session_date__range=date_range,
                submitted = True,
                submitted_at__range = date_range,
                ).count()
        weeks[week_range] = stats
        start_of_week = start_of_week + datetime.timedelta(days=7)
    return weeks

def get_average_survey_ratings(view, semester):
    nms = {
            'rating_1_avg' : "It is clear the ambassador knows and understands the subject matter of this course",
            'rating_2_avg' : "The ambassador explains ideas and concepts clearly.",
            'rating_3_avg' : "The ambassador asks me questions and has me work sample problems.",
            'rating_4_avg' : "The ambassador listens to me and tries to understand my problems.",
            'rating_5_avg' : "The ambassador is friendly and courteous with me.",
            'rating_6_avg' : "The ambassador is trying to accommodate my learning style.",
            'rating_7_avg' : "The session is helpful and improved my understanding of the subject.",
        }
    ambs = TuteeSurvey.objects.all()
    weeks = OrderedDict()
    weeks_since, start_of_week = prepare_range(semester)
    for n in range(weeks_since+1): 
        date_range = (start_of_week, start_of_week + datetime.timedelta(days=6))
        survs_this_week = ambs.filter(
                individual_session__session_date__range=date_range,
                session_canceled=False,
                tutee_absent=False,
                submitted = True,
                )
        week_range = '{:%m/%d/%y} - {:%m/%d/%y}'.format(*date_range)
        stats = OrderedDict()
        avgs = survs_this_week.aggregate(
                rating_1_avg=Avg('rating_1'),
                rating_2_avg=Avg('rating_2'),
                rating_3_avg=Avg('rating_3'),
                rating_4_avg=Avg('rating_4'),
                rating_5_avg=Avg('rating_5'),
                rating_6_avg=Avg('rating_6'),
                rating_7_avg=Avg('rating_7'),
                )
        for k,v in avgs.items():
            stats[nms.get(k)] = v
        weeks[week_range] = stats
        start_of_week = start_of_week + datetime.timedelta(days=7)
    return weeks

def get_missing_tut_surveys():
    missin_tservs = TuteeSurvey.objects.filter(
            individual_session__session_date__lte=timezone.now(),
            submitted=False,).order_by('individual_session__session_date')
    thangs = OrderedDict()
    for tservs in missin_tservs:
        thangs[tservs] = tservs
    return thangs

def get_missing_amb_surveys():
    missin_aservs = AmbassadorSurvey.objects.filter(
            individual_session__session_date__lte=timezone.now(),
            submitted=False,).order_by('individual_session__session_date')
    thangs = OrderedDict()
    for aservs in missin_aservs:
        thangs[aservs] = aservs
    return thangs


def course_stuff():
    everything = Session.objects.all()
    all_courses = Course.objects.all()
    total_sessions = everything.count()
    for course in all_courses:
        course_count = everything.filter(course=course).count()
        # this is really good actually
        if course_count > 0:
            course_percentage = course_count / total_sessions * 100
            # wow...
            yield (course.name, course_count, course_percentage)
        else:
            continue

def tutee_stuff():
    sesshies = IndividualSession.objects.prefetch_related('session').filter(session_date__lte=timezone.now())
    tutees = User.objects.filter(is_staff=False, is_active=True)
    for tut in tutees:
        _pot = sesshies.filter(session__tutee=tut)
        session_count = _pot.count()
        if session_count > 0:
            survs = TuteeSurvey.objects.filter(individual_session__in=_pot, tutee_absent=False, session_canceled=False)
            actual_sessions = survs.filter(submitted=True).count()
            surv_compltn = actual_sessions / survs.count() * 100 if survs.count() > 0 else 0
            courses = [iss.session.course.name for iss in _pot.distinct('session__course')]
            yield (tut.get_full_name(), actual_sessions, session_count, courses, float(f'{surv_compltn:.2f}'))
        else:
            continue

def ambassador_stuff():
    ambs = User.objects.filter(groups__name="Ambassador")
    sesshies = Session.objects.all()
    ambsurv = AmbassadorSurvey.objects.all()
    isesshies = IndividualSession.objects.prefetch_related('session').filter(session_date__lte=timezone.now())
    tsheets = TutoringTimeSheetEntry.objects.all()
    psheets = ProjectTimeSheetEntry.objects.all()
    for amb in ambs:
        _m = sesshies.filter(ambassador=amb)
        _pot = isesshies.filter(session__in=_m)
        s_count = _m.count()
        if s_count == 0:
            s_percent = 0
            t_hours = 0 # this should be 0
        else:
            _as = ambsurv.filter(individual_session__in=_pot)
            s_percent =  _as.filter(submitted=True).count() / _as.count() * 100 if _as.count() > 0 else 0 # forget having it as a string
            t_hours = 0
            for t in tsheets.filter(timesheet__in=amb.timesheets.all()):
                t_hours += t.total_time
        p_hours = 0
        for p in  psheets.filter(project_time_sheet__in=amb.timesheets.all()):
            p_hours += p.total_time
        yield (amb.get_full_name(), s_count, s_percent, t_hours, p_hours)

def add_data(wb):
    ws = wb.active
    ws.title = "Data"
    curr_row = 3 # love those magic numbers... just never change the template file and we're good
    course_generator = course_stuff()
    # sickening
    for tup in course_generator:
        ws.cell(column=1, row=curr_row, value=tup[0])
        ws.cell(column=2, row=curr_row, value=tup[1])
        ws.cell(column=3, row=curr_row, value=tup[2])
        curr_row += 1
    tutee_generator = tutee_stuff()
    # this is so bad lmao
    curr_row = 3
    for tup in tutee_generator:
        ws.cell(column=5, row=curr_row, value=tup[0])
        ws.cell(column=6, row=curr_row, value=tup[1])
        ws.cell(column=7, row=curr_row, value=tup[2])
        ws.cell(column=8, row=curr_row, value='\n'.join(tup[3]))
        ws.cell(column=9, row=curr_row, value=tup[4])
        curr_row += 1

    ambassador_generator = ambassador_stuff()
    curr_row = 3
    for tup in ambassador_generator:
        ws.cell(column=11, row=curr_row, value=tup[0])
        ws.cell(column=12, row=curr_row, value=tup[1])
        ws.cell(column=13, row=curr_row, value=tup[2])
        ws.cell(column=14, row=curr_row, value=tup[3])
        ws.cell(column=15, row=curr_row, value=tup[4])
        curr_row += 1
    return wb

def add_charts(wb):
    total_ambs = User.objects.filter(groups__name="Ambassador").count()
    total_tuts = Session.objects.distinct('tutee').count()
    ws = wb.create_sheet(title="Charts")
    data = wb["Data"]
    # Session Distribution By Course ( Pie )
    pie = PieChart()
    labels = Reference(data, min_col=1, min_row=3, max_row=1000)
    dat = Reference(data, min_col=2, min_row=3, max_row=1000)
    pie.add_data(dat, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Session Distribution By Course"
    ws.add_chart(pie, "B1")

    # Survey Completion By Ambasasdor
    bar1 = BarChart()
    bar1.type = "col"
    bar1.style = 11
    bar1.title = "Survey Completion By Ambassador"
    bar1.x_axis.title = "Ambassador"
    bar1.y_axis.title = "% Of Surveys Completed"
    dat = Reference(data, min_col=13, min_row=3, max_row=2+total_ambs)
    cats = Reference(data, min_col=11, min_row=3, max_row=2+total_ambs)
    bar1.add_data(dat, titles_from_data=False)
    bar1.set_categories(cats)
    bar1.shape = 3
    ws.add_chart(bar1, "L20")

    # Survey Completion By Tutee
    bar1 = BarChart()
    bar1.type = "col"
    bar1.style = 8
    bar1.title = "Survey Completion By Tutee"
    bar1.x_axis.title = "Tutee"
    bar1.y_axis.title = "% Of Surveys Completed"
    dat = Reference(data, min_col=9, min_row=3, max_row=2+total_tuts)
    cats = Reference(data, min_col=5, min_row=3, max_row=2+total_tuts)
    bar1.add_data(dat, titles_from_data=False)
    bar1.set_categories(cats)
    bar1.shape = 2
    ws.add_chart(bar1, "V1")

    # Total Sessions By Ambassador
    bar1 = BarChart()
    bar1.type = "col"
    bar1.style = 11
    bar1.title = "Total Sessions By Ambassadors"
    bar1.x_axis.title = "Ambassador"
    bar1.y_axis.title = "# Of Sessions"
    dat = Reference(data, min_col=12, min_row=3, max_row=2+total_ambs)
    cats = Reference(data, min_col=11, min_row=3, max_row=2+total_ambs)
    bar1.add_data(dat, titles_from_data=False)
    bar1.set_categories(cats)
    bar1.shape = 3
    ws.add_chart(bar1, "L1")

    # Total Session vs Actual Sesssion
    bar1 = BarChart()
    bar1.type = "bar"
    bar1.style = 10
    bar1.title = "Expected Sessions Vs Actual Sessions"
    bar1.y_axis.title = "# Of Sessions"

    dat = Reference(data, min_col=18, min_row=2, max_row=5)
    cats = Reference(data, min_col=17, min_row=2, max_row=5)
    bar1.add_data(dat, titles_from_data=False)
    bar1.set_categories(cats)
    bar1.shape = 4
    ws.add_chart(bar1, "B20")
    return wb
