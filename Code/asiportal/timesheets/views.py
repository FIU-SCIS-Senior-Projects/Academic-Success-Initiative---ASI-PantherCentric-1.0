from itertools import chain
from django.db.models import Sum
from django.contrib.auth.models import User
from django.core.urlresolvers import reverse_lazy
from django.shortcuts import render
from django.views import generic
from django.utils import timezone
from django.http import HttpResponse, HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from projects.models import ProjectTS, ProjectTSEntry
from courses.models import Course
from surveys.models import AmbassadorSurvey, TuteeSurvey
from tutoring_sessions.models import IndividualSession 
from .models import TimeSheet, TutoringTimeSheetEntry
from asiapp.mixins import (
        LoginRequiredMixin,
        ProjectPermRequired,
        AmbassadorPermRequired)
from .forms import (
        TimeSheetEntryFormSet,
        TimeSheetForm, 
        TimeSheetEntryApprovalFormSet,
        FinalApproveTimeSheetForm,
        )

def download_timesheet(request, ambassador, pk):
    if request.user.username == ambassador or request.user.has_perm('timesheets.approve_timesheet'):
        amb = User.objects.get(username=ambassador)
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        timesheet = TimeSheet.objects.get(pk=pk)
        _entries = TutoringTimeSheetEntry.objects.prefetch_related('session', 'session__session').filter(timesheet=timesheet).order_by('session__session_date', 'session__session__start_time')
        _ind = [ent.session for ent in _entries]
        _amb_survs = AmbassadorSurvey.objects.filter(individual_session__in=_ind)
        _tut_survs = TuteeSurvey.objects.filter(individual_session__in=_ind)
        response = HttpResponse(content_type='application/ms-excel')
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename={}-hours-{}-{}.xlsx'\
                .format(amb.first_name, amb.last_name, timesheet.pay_period_begin, timesheet.pay_period_end)
        wb = load_workbook(
                'tutoring_sessions/templates/tutoring_sessions/timesheet_template.xlsx')
        ws = wb.active
        ws.title = '{} - {}'.format(timesheet.pay_period_begin, timesheet.pay_period_end)
        ws['A1'].value="Ambassador Name: {}".format(timesheet.ambassador.get_full_name())
        ws2 = wb.copy_worksheet(ws)
        ws2.title = 'Projects'
        for row, entry in zip(range(3, _entries.count()+3), _entries):
            asurv = _amb_survs.get(individual_session = entry.session)
            tsurv = _tut_survs.get(individual_session = entry.session)
            ws.cell(column=1, row=row, value="Y" if asurv.submitted else "N")
            ws.cell(column=2, row=row, value="Y" if tsurv.submitted else "N")
            ws.cell(column=3, row=row, value="{}".format(entry.session.session_date))  # date
            ws.cell(column=4, row=row, value="{}".format(entry.session.session.start_time.strftime("%-I:%M %p"))) # start time
            ws.cell(column=5, row=row, value="{}".format(entry.session.session.end_time.strftime("%-I:%M %p"))) # end time
            ws.cell(column=6, row=row, value=entry.total_time) # 
            ws.cell(column=7, row=row, value=entry.session.session.course.__str__()) # session
            ws.cell(column=8, row=row, value="{}".format(entry.session.session.tutee.get_full_name())) # tutee
            ws.cell(column=10, row=row, value="{}".format(entry.session.session.room_number)) # room #
            ws.cell(column=11, row=row, value="{}".format(asurv.comments)) # room #
            ws.cell(column=12, row=row, value="{}".format("Yes" if entry.tl_verified else "No")) # room #
        img = Image('tutoring_sessions/templates/tutoring_sessions/asiclear.png')
        ws.add_image(img, 'E9')
        try:
            ptimesheet = ProjectTS.objects.get(pay_period_begin__gte=timesheet.pay_period_begin,
                    pay_period_end__lte=timesheet.pay_period_end,
                    ambassador=amb,
                    )
            _projent = ProjectTSEntry.objects.filter(project_time_sheet=ptimesheet).order_by('day', 'start_time')
            ws = ws2
            for row, entry in zip(range(3, _projent.count()+3), _projent):
                ws.cell(column=1, row=row, value="-")
                ws.cell(column=2, row=row, value="-")
                ws.cell(column=3, row=row, value="{}".format(entry.day))  # date
                ws.cell(column=4, row=row, value="{}".format(entry.start_time.strftime("%-I:%M %p"))) # start time
                ws.cell(column=5, row=row, value="{}".format(entry.end_time.strftime("%-I:%M %p"))) # end time
                ws.cell(column=6, row=row, value=entry.total_time) # 
                ws.cell(column=7, row=row, value=entry.title) # project
                ws.cell(column=8, row=row, value="-") # tutee
                ws.cell(column=10, row=row, value="-") # room #
                ws.cell(column=11, row=row, value="{}".format(entry.description)) 
                ws.cell(column=12, row=row, value="{}".format("Yes" if entry.project_leader_verification else "No"))
            img = Image('tutoring_sessions/templates/tutoring_sessions/asiclear.png')
            ws.add_image(img, 'E9')
            wb.save(response)
            return response
        except ObjectDoesNotExist:
            wb.save(response)
            return response
    else:
        return HttpResponse('')


def approve_timesheet(request):
    if request.method == 'GET' and request.user.has_perm('timesheets.approve_entry'):
        # Q U E  R Y  O P T I M I Z A T I O N for bad DB design
        _all = TutoringTimeSheetEntry.objects.prefetch_related(
                'session__session',
                'session',
                'session__session__course',
                'session__session__tutee',
                'session__session__ambassador').all()
        _is = IndividualSession.objects.prefetch_related('session').all()
        _ours = Course.objects.filter(team_leader=request.user)
        _sessions = _is.filter(session__course__in=_ours)
        _final = _all.filter(session__in = _sessions, tl_verified=False)
        formset = TimeSheetEntryApprovalFormSet(queryset = _final)
        return render(request,
                'timesheets/timesheet_approve.html',
                {'formset' : formset},
            )
    elif request.method == 'POST':
        formset = TimeSheetEntryApprovalFormSet(request.POST)
        formset.save()
        return HttpResponseRedirect(reverse_lazy('timesheets:timesheet_list'))
    else:
        return HttpResponseRedirect(reverse_lazy('timesheets:timesheet_list'))


def edit_timesheet(request, timesheet_pk):
    papa = TimeSheet.objects.get(pk=timesheet_pk)
    if request.user != papa.ambassador:
        return HttpResponseRedirect(reverse_lazy('timesheets:timesheet_list'))
    formset = TimeSheetEntryFormSet(form_kwargs={'user' : request.user, 'parent' : papa})
    entries = TutoringTimeSheetEntry.objects.prefetch_related('session',
            'session__session__ambassador',
            'session__session__tutee',
            'session__session').filter(timesheet = papa)
    if request.method == 'GET':
        return render(request,
                'timesheets/tutoringtimesheet_form.html',
                {'formset' : formset,
                 'timesheet' : papa,
                'entries' : entries},
                )

class TimeSheetDetailView(LoginRequiredMixin, generic.DetailView):
    model = TimeSheet
    pk_url_kwarg = 'timesheet_pk'

    def dispatch(self, request, *args, **kwargs):
        ts = TimeSheet.objects.get(pk=kwargs.get('timesheet_pk'))
        if ts.ambassador != request.user:
            return HttpResponseRedirect(reverse_lazy('timesheets:timesheet_list'))
        else:
            return super(TimeSheetDetailView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, *args, **kwargs):
        cd = super(TimeSheetDetailView, self).get_context_data(*args, **kwargs)
        _all = TutoringTimeSheetEntry.objects.prefetch_related('session', 'session__session').filter(timesheet=self.kwargs.get('timesheet_pk')).order_by('session__session_date', 'session__session__start_time')
        cd['entries'] = _all
        cd['total_time'] = _all.aggregate(Sum('total_time')).get('total_time__sum')
        return cd

class TimeSheetListView(LoginRequiredMixin, generic.ListView):
    model = TimeSheet

    def get_context_data(self, **kwargs):
        allSheets = TimeSheet.objects.filter(ambassador=self.request.user)
        can_approve = self.request.user.has_perm('timesheets.approve_entry')

        cd = {
            'incomplete' : allSheets.filter(final_approval=False),
            'can_approve' : can_approve,
            'approved' : allSheets.filter(
                final_approval = True,
                )
            }
        return cd


# needs protection
class TimeSheetCreateFormView(AmbassadorPermRequired, generic.CreateView):
    form_class = TimeSheetForm
    model = TimeSheet
    success_url = reverse_lazy('timesheets:timesheet_list')

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.ambassador = self.request.user
        self.object.save()
        return super(TimeSheetCreateFormView, self).form_valid(form)

# needs protection
class AmbassadorListView(ProjectPermRequired, generic.ListView):
    model = User
    template_name = 'timesheets/amb_dir.html'

    def get_queryset(self):
        qs = User.objects.filter(groups__name='Ambassador')
        return qs

class TimeSheetAmbassadorListView(ProjectPermRequired, generic.ListView):
    model = TimeSheet
    template_name = 'timesheets/ambassador_timesheet_list.html'

    def get_context_data(self, *args, **kwargs):
        cd = super(TimeSheetAmbassadorListView, self).get_context_data(*args, **kwargs)
        cd['ambassador'] = self.kwargs.get('ambassador')
        return cd

    def get_queryset(self):
        qs = TimeSheet.objects.filter(ambassador__username=self.kwargs.get('ambassador')).order_by('final_approval')
        return qs


class FinalApproveAmbassadorTimeSheet(ProjectPermRequired, generic.UpdateView):
    model = TimeSheet
    form_class = FinalApproveTimeSheetForm
    template_name = 'timesheets/final_approve_form.html'
    success_url = reverse_lazy('timesheets:ambassador_list')

    def get_context_data(self, *args, **kwargs):
        cd = super(FinalApproveAmbassadorTimeSheet, self).get_context_data(*args, **kwargs)
        _all = TutoringTimeSheetEntry.objects.prefetch_related('session', 'session__session').filter(timesheet=self.kwargs.get('pk')).order_by('session__session_date', 'session__session__start_time')
        cd['related_sheets'] = _all
        cd['total_time'] = _all.aggregate(Sum('total_time'))['total_time__sum']
        return cd

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.final_approval = True
        self.object.save()
        return HttpResponseRedirect(reverse_lazy('timesheets:ambassador_sheets', kwargs={'ambassador' : self.object.ambassador.username}))
